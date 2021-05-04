import cv2
import pickle
# import xlwt
import openpyxl
from openpyxl import Workbook
# import xlrd
import datetime
# from xlwt import Workbook
import numpy as np
parth = 0
niraj = 0
def face():
    cp = 0
    cn = 0
    label = {"person": 1}
    with open("labels.pickle", "rb") as f:
        label_ = pickle.load(f)
        label = {v: k for k, v in label_.items()}
    face = cv2.CascadeClassifier("haarcascade_frontalface_default.xml")
    img = cv2.VideoCapture(0)
    rec = cv2.face.LBPHFaceRecognizer_create()
    rec.read("train.yml")


    while True:
        rat, frame = img.read()
        gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
        faces = face.detectMultiScale(gray, scaleFactor=1.5, minNeighbors=5)
        for (x, y, w, h) in faces:
            roi_gray = gray[y:y+h, x:x+w]
            roi_color = frame[y:y+h, x:x+w]

            id_, con = rec.predict(roi_gray)
            if con >= 45:
                print(id_)
                if "niraj" ==label[id_]:
                    cn= cn+1
                if "parth" ==label[id_]:
                    cp =cp+1
                print(label[id_])
                cv2.rectangle(frame, (x, y), (x+w, y+h), (0, 255, 0))
            if cp>45:
                global parth
                parth=1
            if cn>45:
                global niraj
                niraj=1
            cv2.putText(frame, label[id_], (x, y), cv2.FONT_HERSHEY_COMPLEX, 1, (255, 255, 255), 2, cv2.LINE_AA)

        cv2.imshow("Face", frame)

        if cv2.waitKey(1) & 0xFF == ord("q"):
            break
    cv2.destroyAllWindows()
    img.release()


    # date = datetime.datetime.now()
    # a = date.date()
    #
    # wb = openpyxl.load_workbook('book1.xlsx')
    # sheet = wb['Sheet1']
    # # b = sheet[('B'+1)+'2'].value.date()
    # data2 = 1
    #
    # # Workbook is created
    #
    #
    # # add_sheet is used to create sheet.
    # for i in range(2,20):
    #     if(a==sheet.cell(row = 9, column = i).value.date()):
    #         data2 = i
    #         break
    #     else:
    #         pass
    #
    # if(bhadresh!=0):
    #     sheet.cell(row=10, column=data2).value = 1
    # else:
    #     sheet.cell(row=10, column=i).value = 0
    # if(parth!=0):
    #     sheet.cell(row=11, column=i).value = 1
    # else:
    #     sheet.cell(row=11, column=i).value = 0
    # wb.save('Book1.xlsx')
if __name__ =="__main__":
    face()