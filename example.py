# import pandas as pd
# from openpyxl import load_workbook
#
# df2 = pd.DataFrame({'Data': [13, 24, 35, 46]})
# book = load_workbook('book2.xlsx')
# writer = pd.ExcelWriter('book2.xlsx', engine='openpyxl')
# df2["jrry"] =[45,85,63,1]
# df2.to_excel(writer, "Sheet1", startcol=10,startrow=10)
# writer.save()

import openpyxl
wb = openpyxl.load_workbook('book2.xlsx')
wb2 = openpyxl.load_workbook('book3.xlsx')
sheet = wb['Sheet1']
sheet2 = wb2['Sheet1']


rows = 6
columns = 4

listtab = []

for i in range(1,rows+1):
    listtab.append([])

for r in range(1,rows+1):
    for c in range(1,columns+1):
        e= sheet.cell(row=r,column=c)
        listtab[r-1].append(e.value)
print(listtab)
for r in range(1,rows+1):
    for c in range(1,columns+1):
        j= sheet2.cell(row=r,column=c)
        j.value = listtab[r-1][c-1]

wb2.save('book3.xlsx')